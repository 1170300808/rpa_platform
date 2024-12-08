from openpyxl import load_workbook, Workbook
from tkinter import simpledialog, filedialog, messagebox
from copy_cell_styles import copy_cell_styles


def run(file_list):
    if not file_list:
        messagebox.showerror("错误", "未上传任何文件")
        return

    try:
        # 1. 用户输入保留的前 n 行
        n = simpledialog.askinteger("输入", "请输入需要保留的前 n 行：", minvalue=1)
        if n is None:
            return  # 用户取消输入时直接退出

        # 2. 初始化变量
        common_header = []  # 用于存储第一个文件的前 n 行
        combined_data = []  # 用于存储从每个文件 n+1 行开始的数据

        content_size = []
        # 3. 遍历所有已选择的文件
        for file in file_list:
            workbook = load_workbook(file)
            sheet = workbook.active

            # (a) 读取前 n 行作为公共头部
            current_header = []
            for i in range(1, n + 1):  # openpyxl 索引从 1 开始
                current_header.append([cell.value for cell in sheet[i]])

            # 初始化 common_header，如果已有值则比较一致性
            if not common_header:
                common_header = current_header
            elif common_header != current_header:
                messagebox.showerror("错误", f"文件 {file} 的前 {n} 行与其他文件不一致！")
                return

            # (b) 读取从第 n+1 行开始的数据
            size = 0
            for row in sheet.iter_rows(min_row=n + 1, values_only=True):
                combined_data.append(row)
                size += 1
            content_size.append(size)

        # 4. 创建新工作簿并写入数据
        new_workbook = Workbook()
        new_sheet = new_workbook.active

        # (a) 写入公共头部
        for header_row in common_header:
            new_sheet.append(header_row)

        # (b) 写入合并数据
        for data_row in combined_data:
            new_sheet.append(data_row)

        # 5. 保存最终合并的文件
        save_path = filedialog.asksaveasfilename(
            title="保存合并后的文件",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")]
        )
        if save_path:
            new_workbook.save(save_path)
            messagebox.showinfo("成功", "文件合并成功并已保存！")

        # 先复制表头样式
        min_row, max_row, min_col, max_col = 1, len(common_header), 1, len(common_header[0])
        size = (min_row, max_row, min_col, max_col)
        copy_cell_styles(file_list[0], save_path, size, size)

        # 再复制内容样式
        for idx, file in enumerate(file_list):
            min_row = max_row + 1
            max_row = min_row + content_size[idx] - 1
            size = (min_row, max_row, min_col, max_col)
            copy_cell_styles(file, save_path,
                             (1 + len(common_header), len(common_header) + content_size[idx], min_col, max_col),
                             size)

    except Exception as e:
        messagebox.showerror("错误", f"合并文件失败：{e}")
