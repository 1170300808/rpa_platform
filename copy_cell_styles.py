from copy import copy
from tkinter import messagebox, Toplevel, Label, Entry, Button, simpledialog
import re

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries


def run(file_list):
    # if not file_list:
    #     messagebox.showerror("错误", "未上传任何文件")
    #     return
    if not file_list or len(file_list) != 2:
        messagebox.showerror("错误", "该功能需上传两个文件，第一个文件为源文件，第二个文件为目标文件")
        return
    try:
        # 弹窗输入两个字符串
        input1 = simpledialog.askstring("输入源文件单元格范围", "输入源文件单元格范围：")

        if not is_valid_cell_range(input1):
            messagebox.showerror("错误", "单元格范围错误")
            return

        input2 = simpledialog.askstring("输入目标文件单元格范围", "输入目标文件单元格范围：")
        if not is_valid_cell_range(input2):
            messagebox.showerror("错误", "单元格范围错误")
            return

        # 显示输入的结果
        # messagebox.showinfo("输入结果", f"第一个字符串：{input1}\n第二个字符串：{input2}")
        copy_cell_styles(file_list[0], file_list[1], input1, input2)
    except Exception as e:
        messagebox.showerror("错误", f"合并文件失败：{e}")


def is_valid_cell_range(range_string):
    # 正则表达式：匹配 A1:B2 格式的范围
    pattern = r'^([A-Z]+)(\d+):([A-Z]+)(\d+)$'

    match = re.match(pattern, range_string)

    if match:
        start_col = match.group(1)  # 起始列字母
        start_row = int(match.group(2))  # 起始行数字
        end_col = match.group(3)  # 结束列字母
        end_row = int(match.group(4))  # 结束行数字

        return start_col <= end_col and start_row <= end_row

    return False


def copy_cell_styles(source_file, target_file, source_range, target_range):
    # 加载源文件和目标文件
    source_wb = load_workbook(source_file)
    target_wb = load_workbook(target_file)

    # 获取工作表（假设使用第一个工作表）
    source_ws = source_wb.active
    target_ws = target_wb.active

    # 获取源和目标区域的行列范围
    if isinstance(source_range, tuple):
        source_cells = list(source_ws.iter_rows(
            min_row=source_range[0], max_row=source_range[1],
            min_col=source_range[2], max_col=source_range[3]
        ))
        target_cells = list(target_ws.iter_rows(
            min_row=target_range[0], max_row=target_range[1],
            min_col=target_range[2], max_col=target_range[3]
        ))
    else:
        source_cells = source_ws[source_range]
        target_cells = target_ws[target_range]
        if len(source_cells) != len(target_cells) or len(source_cells[0]) != len(target_cells[0]):
            raise ValueError("源区域和目标区域的大小不匹配")

    for merged_range in source_ws.merged_cells.ranges:
        print(merged_range)

    # 确保源和目标区域的大小一致

    copy_merged_cells(source_ws, target_ws, source_range, target_range)

    # 遍历源区域的单元格并复制样式到目标区域
    for i, row in enumerate(source_cells):
        for j, cell in enumerate(row):
            target_cell = target_cells[i][j]

            # 复制内容
            target_cell.value = cell.value

            # 复制字体
            target_cell.font = copy(cell.font)
            # 复制对齐方式
            target_cell.alignment = copy(cell.alignment)
            # 复制边框
            target_cell.border = copy(cell.border)
            # 复制填充
            target_cell.fill = copy(cell.fill)
            # 复制数字格式
            target_cell.number_format = copy(cell.number_format)
            # # 复制超链接
            # target_cell.hyperlink = cell.hyperlink
            # # 复制保护状态
            # target_cell.protection = cell.protection

    # 保存目标文件
    target_wb.save(target_file)


def copy_merged_cells(source_sheet, target_sheet, source_range, target_range):
    # 获取源范围和目标范围的边界
    if isinstance(source_range, str):
        min_col_s, min_row_s, max_col_s, max_row_s = range_boundaries(source_range)
        min_col_t, min_row_t, max_col_t, max_row_t = range_boundaries(target_range)
    else:
        min_row_s, max_row_s, min_col_s, max_col_s = source_range
        min_row_t, max_row_t, min_col_t, max_col_t = target_range

    # 确保源范围和目标范围大小一致
    if (max_row_s - min_row_s) != (max_row_t - min_row_t) or (max_col_s - min_col_s) != (max_col_t - min_col_t):
        raise ValueError("源范围和目标范围的大小不一致")

    # 获取源范围的合并单元格
    for merged_range in source_sheet.merged_cells.ranges:
        # 判断合并单元格是否在源范围内
        min_col_m, min_row_m, max_col_m, max_row_m = range_boundaries(str(merged_range))
        if min_col_m >= min_col_s and max_col_m <= max_col_s and min_row_m >= min_row_s and max_row_m <= max_row_s:
            # 计算在目标范围中的对应位置
            offset_col = min_col_m - min_col_s
            offset_row = min_row_m - min_row_s
            target_min_col = min_col_t + offset_col
            target_min_row = min_row_t + offset_row
            target_max_col = target_min_col + (max_col_m - min_col_m)
            target_max_row = target_min_row + (max_row_m - min_row_m)

            # 合并目标单元格
            target_sheet.merge_cells(start_row=target_min_row, start_column=target_min_col,
                                     end_row=target_max_row, end_column=target_max_col)

# 使用示例
# copy_cell_styles('source.xlsx', 'target.xlsx', 'A1:B2', 'C3:D4')

# print(is_valid_cell_range('A1:C3'))
